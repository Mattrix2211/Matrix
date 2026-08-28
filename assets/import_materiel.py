"""Import en masse de matériel mobile (Asset) depuis un fichier Excel.

Périmètre volontairement restreint à la première version de cette fonctionnalité
(cf. tâche Notion Phase 6) : uniquement le matériel mobile (`Asset`), pas les
installations fixes. Réutilise openpyxl en lecture (déjà présent dans le projet
pour l'export, cf. matrix/core/export.py) plutôt que d'ajouter une nouvelle
dépendance, et réutilise le contrôle de périmètre déjà en place pour la création
manuelle d'un matériel (`_org_dans_perimetre`, assets/web_views.py).

Import atomique PAR LIGNE : une ligne en erreur (colonne manquante, type
inconnu, numéro de série déjà utilisé...) n'empêche pas les autres lignes
valides du même fichier d'être importées. Chaque ligne en erreur est reportée
avec son numéro et un message en français, pour correction dans le tableur
puis un nouvel import (pas de système de brouillon à valider : le fichier est
traité directement, conformément à la demande).
"""
import zipfile
from dataclasses import dataclass, field

from django.core.exceptions import ValidationError
from django.db import transaction

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from matrix.core.export import rendre_xlsx_sections
from org.models import Ship, Service, Sector, Section
from accounts.models import AuditLog
from .models import Asset, AssetType, Location

# En-têtes du fichier Excel attendu, dans cet ordre. Les six premières colonnes
# reprennent exactement les libellés déjà utilisés par l'export de la liste de
# matériel (_ENTETES_EXPORT_ASSETS, assets/web_views.py) pour rester cohérent
# avec un format déjà connu des utilisateurs ; les colonnes suivantes ajoutent
# les champs nécessaires à la création qui ne figurent pas dans l'export.
ENTETES_MODELE = [
    'Désignation', 'Type', 'Identifiant interne', 'N° série', 'Statut', 'Criticité',
    'Unité', 'Service', 'Secteur', 'Section', 'Emplacement',
    'Référence', 'Marque', 'NNO', 'Gisement', 'Local',
]

# Colonnes obligatoires : sans elles, le fichier est rejeté avant même de lire
# les lignes (message d'erreur générique, pas de rapport ligne par ligne).
_COLONNES_OBLIGATOIRES = {'désignation': 'Désignation', 'type': 'Type'}

_LIGNE_EXEMPLE = [
    'Extincteur CO2 6 kg', 'Extincteur', 'EXT-0001', 'SN-123456', 'OK', 1,
    '', '', '', '', 'Local machines',
    'REF-001', 'Marque X', 'NNO-001', 'Gisement 3', 'Local 12',
]

_INSTRUCTIONS = [
    ('Désignation', "Obligatoire. Nom du matériel (ex : Extincteur CO2 6 kg)."),
    ('Type', "Obligatoire. Doit correspondre exactement (accents/majuscules non "
             "significatifs) au nom d'un type de matériel déjà créé dans le secteur visé."),
    ('Identifiant interne', "Facultatif."),
    ('N° série', "Facultatif. Doit être unique : une ligne est refusée si ce numéro "
                 "est déjà utilisé par un autre matériel."),
    ('Statut', "Facultatif (OK par défaut). Valeurs possibles : OK, En service, Hors "
               "service, Défectueux."),
    ('Criticité', "Facultatif (1 par défaut). Nombre entier de 1 à 5."),
    ('Unité', "Facultatif si votre profil est déjà rattaché à une unité. Sinon, nom "
              "exact de l'unité (navire, école...)."),
    ('Service', "Facultatif si votre profil est déjà rattaché à un service. Sinon, "
                "nom exact du service."),
    ('Secteur', "Obligatoire si votre profil n'est pas déjà rattaché à un secteur. "
                "Nom exact du secteur."),
    ('Section', "Facultatif. Nom exact de la section."),
    ('Emplacement', "Facultatif. Créé automatiquement dans l'unité si le nom n'existe pas encore."),
    ('Référence', "Facultatif."),
    ('Marque', "Facultatif."),
    ('NNO', "Facultatif."),
    ('Gisement', "Facultatif."),
    ('Local', "Facultatif."),
]

# Statuts acceptés dans la colonne "Statut" : code technique et libellé français
# affiché, tous deux acceptés sans tenir compte de la casse (cf. Asset.STATUS).
_STATUTS_PAR_LIBELLE = {code.lower(): code for code, _libelle in Asset.STATUS}
_STATUTS_PAR_LIBELLE.update({libelle.lower(): code for code, libelle in Asset.STATUS})


class _ErreurLigneImport(Exception):
    """Erreur métier propre à une ligne du fichier importé (message déjà en
    français, prêt à être reporté tel quel dans le rapport)."""


@dataclass
class ResultatImportMateriel:
    """Rapport d'un import : nombre de lignes traitées, nombre de matériels
    effectivement créés, et messages d'erreur (un par ligne en échec, ou une
    erreur générale si le fichier lui-même est illisible/mal formé)."""
    total_lignes: int = 0
    crees: int = 0
    erreurs: list = field(default_factory=list)

    @property
    def a_des_erreurs(self) -> bool:
        return bool(self.erreurs)


def generer_modele_xlsx():
    """Construit le classeur Excel téléchargeable documentant le format attendu
    (feuille "Modèle" avec en-têtes + ligne d'exemple, feuille "Instructions"
    avec le détail de chaque colonne). Renvoie None si openpyxl n'est pas
    disponible (même garde-fou que le reste des exports, cf. matrix/core/export.py)."""
    sections = [
        ("Modèle", ENTETES_MODELE, [_LIGNE_EXEMPLE]),
        ("Instructions", ["Colonne", "Explication"], _INSTRUCTIONS),
    ]
    return rendre_xlsx_sections(sections)


def _normaliser(texte) -> str:
    return (texte or '').strip().lower()


def _valeur_colonne(ligne, index_colonnes, nom_normalise):
    """Lit la valeur d'une colonne dans une ligne du tableur, par nom d'en-tête
    déjà normalisé. Renvoie une chaîne vide si la colonne est absente ou vide."""
    idx = index_colonnes.get(nom_normalise)
    if idx is None or idx >= len(ligne):
        return ''
    valeur = ligne[idx]
    if valeur is None:
        return ''
    if isinstance(valeur, float) and valeur.is_integer():
        # Excel stocke souvent les nombres entiers saisis en flottant (1 -> 1.0) ;
        # évite d'afficher/comparer "1.0" au lieu de "1" (ex: criticité).
        valeur = int(valeur)
    return str(valeur).strip()


def _ligne_vide(ligne) -> bool:
    return ligne is None or all(c is None or str(c).strip() == '' for c in ligne)


def _resoudre_organisation(ligne, index_colonnes, user, profil):
    """Résout l'unité/service/secteur/section d'une ligne à partir des colonnes
    Unité/Service/Secteur/Section (par nom, pas par identifiant technique — le
    fichier reste lisible/remplissable par un utilisateur métier) et, à défaut,
    du périmètre déjà affecté au profil de l'utilisateur (même logique que
    _perimetre_utilisateur pour le formulaire de création manuelle). Contrôle
    ensuite que chaque valeur explicitement fournie appartient bien au périmètre
    de l'appelant, en réutilisant _org_dans_perimetre (assets/web_views.py) —
    import différé ci-dessous pour éviter un import circulaire avec ce module."""
    from .web_views import _org_dans_perimetre  # import différé, voir docstring

    nom_unite = _valeur_colonne(ligne, index_colonnes, 'unité')
    nom_service = _valeur_colonne(ligne, index_colonnes, 'service')
    nom_secteur = _valeur_colonne(ligne, index_colonnes, 'secteur')
    nom_section = _valeur_colonne(ligne, index_colonnes, 'section')

    ship = None
    if nom_unite:
        ship = Ship.objects.filter(name__iexact=nom_unite).first()
        if ship is None:
            raise _ErreurLigneImport(f"unité inconnue : « {nom_unite} »")
        if not _org_dans_perimetre(user, Ship, ship.id):
            raise _ErreurLigneImport(f"unité hors de votre périmètre : « {nom_unite} »")
    else:
        ship = getattr(profil, 'ship', None)

    service = None
    if nom_service:
        qs = Service.objects.filter(name__iexact=nom_service)
        if ship is not None:
            qs = qs.filter(ship=ship)
        service = qs.first()
        if service is None:
            raise _ErreurLigneImport(f"service inconnu : « {nom_service} »")
        if not _org_dans_perimetre(user, Service, service.id):
            raise _ErreurLigneImport(f"service hors de votre périmètre : « {nom_service} »")
        ship = ship or service.ship
    else:
        service = getattr(profil, 'service', None)
        if service is not None:
            ship = ship or service.ship

    sector = None
    if nom_secteur:
        qs = Sector.objects.filter(name__iexact=nom_secteur)
        if service is not None:
            qs = qs.filter(service=service)
        sector = qs.first()
        if sector is None:
            raise _ErreurLigneImport(f"secteur inconnu : « {nom_secteur} »")
        if not _org_dans_perimetre(user, Sector, sector.id):
            raise _ErreurLigneImport(f"secteur hors de votre périmètre : « {nom_secteur} »")
        service = service or sector.service
        ship = ship or sector.service.ship
    else:
        sector = getattr(profil, 'sector', None)
        if sector is not None:
            service = service or sector.service
            ship = ship or sector.service.ship

    if sector is None:
        raise _ErreurLigneImport(
            "secteur manquant : renseignez la colonne Secteur (aucun secteur par défaut sur votre profil)"
        )
    if ship is None or service is None:
        raise _ErreurLigneImport(
            "unité ou service manquant : renseignez les colonnes Unité et Service"
        )

    section = None
    if nom_section:
        section = Section.objects.filter(name__iexact=nom_section, sector=sector).first()
        if section is None:
            raise _ErreurLigneImport(
                f"section inconnue : « {nom_section} » pour le secteur « {sector.name} »"
            )
        if not _org_dans_perimetre(user, Section, section.id):
            raise _ErreurLigneImport(f"section hors de votre périmètre : « {nom_section} »")
    else:
        profil_section = getattr(profil, 'section', None)
        if profil_section is not None and profil_section.sector_id == sector.id:
            section = profil_section

    return ship, service, sector, section


def _resoudre_statut(ligne, index_colonnes):
    brut = _valeur_colonne(ligne, index_colonnes, 'statut')
    if not brut:
        return 'OK'
    code = _STATUTS_PAR_LIBELLE.get(brut.lower())
    if code is None:
        raise _ErreurLigneImport(
            f"statut inconnu : « {brut} » (valeurs possibles : OK, En service, Hors service, Défectueux)"
        )
    return code


def _resoudre_criticite(ligne, index_colonnes):
    brut = _valeur_colonne(ligne, index_colonnes, 'criticité')
    if not brut:
        return 1
    try:
        valeur = int(brut)
    except ValueError:
        raise _ErreurLigneImport(f"criticité invalide : « {brut} » (attendu un nombre entier de 1 à 5)")
    if not (1 <= valeur <= 5):
        raise _ErreurLigneImport(f"criticité invalide : « {brut} » (attendu un nombre entier de 1 à 5)")
    return valeur


def _creer_asset_depuis_ligne(ligne, index_colonnes, user, profil, numeros_serie_vus):
    """Valide et crée un `Asset` à partir d'une ligne du tableur. Lève
    `_ErreurLigneImport` (message métier déjà en français) ou `ValidationError`
    (via full_clean, ex : protection anti-cycle) en cas d'échec — dans les deux
    cas, rien n'est enregistré (appelé sous transaction.atomic() par l'appelant)."""
    designation = _valeur_colonne(ligne, index_colonnes, 'désignation')
    if not designation:
        raise _ErreurLigneImport("désignation obligatoire manquante")
    nom_type = _valeur_colonne(ligne, index_colonnes, 'type')
    if not nom_type:
        raise _ErreurLigneImport("type de matériel obligatoire manquant")

    ship, service, sector, section = _resoudre_organisation(ligne, index_colonnes, user, profil)

    asset_type = AssetType.objects.filter(sector=sector, name__iexact=nom_type).first()
    if asset_type is None:
        raise _ErreurLigneImport(f"type de matériel inconnu : « {nom_type} » (secteur {sector.name})")

    statut = _resoudre_statut(ligne, index_colonnes)
    criticite = _resoudre_criticite(ligne, index_colonnes)

    serial = _valeur_colonne(ligne, index_colonnes, 'n° série')
    if serial and (serial in numeros_serie_vus or Asset.objects.filter(serial_number=serial).exists()):
        raise _ErreurLigneImport(f"numéro de série déjà utilisé : « {serial} »")

    nom_emplacement = _valeur_colonne(ligne, index_colonnes, 'emplacement')
    location = None
    if nom_emplacement:
        location, _cree = Location.objects.get_or_create(ship=ship, name=nom_emplacement, parent=None)

    asset = Asset(
        asset_type=asset_type,
        designation=designation,
        internal_id=_valeur_colonne(ligne, index_colonnes, 'identifiant interne'),
        serial_number=serial,
        reference=_valeur_colonne(ligne, index_colonnes, 'référence'),
        marque=_valeur_colonne(ligne, index_colonnes, 'marque'),
        nno=_valeur_colonne(ligne, index_colonnes, 'nno'),
        gisement=_valeur_colonne(ligne, index_colonnes, 'gisement'),
        local=_valeur_colonne(ligne, index_colonnes, 'local'),
        status=statut,
        criticality=criticite,
        ship=ship,
        service=service,
        sector=sector,
        section=section,
        location=location,
    )
    asset.full_clean()
    asset.save()
    if serial:
        numeros_serie_vus.add(serial)
    AuditLog.objects.create(
        actor=user, action='import_asset',
        details=f'type_id={asset_type.id}; internal_id={asset.internal_id}',
    )


def importer_materiel_depuis_fichier(fichier, user) -> ResultatImportMateriel:
    """Point d'entrée de l'import : lit le fichier Excel uploadé, valide les
    en-têtes, puis traite chaque ligne indépendamment (import atomique par
    ligne — une ligne en erreur n'empêche pas les suivantes d'être importées)."""
    try:
        classeur = load_workbook(fichier, data_only=True, read_only=True)
        feuille = classeur.worksheets[0]
        lignes = list(feuille.iter_rows(values_only=True))
    except (InvalidFileException, zipfile.BadZipFile, OSError, KeyError, ValueError, IndexError):
        # Capture volontairement large (openpyxl peut lever plusieurs types
        # d'exceptions selon la corruption du fichier) : un fichier illisible ne
        # doit jamais faire planter la page, seulement afficher un message clair
        # invitant à repartir du modèle fourni.
        return ResultatImportMateriel(
            erreurs=["Fichier illisible : utilisez un fichier Excel (.xlsx) valide, "
                     "de préférence téléchargé depuis le modèle fourni."]
        )

    if not lignes:
        return ResultatImportMateriel(erreurs=["Le fichier est vide : aucune ligne à importer."])

    index_colonnes = {_normaliser(valeur): i for i, valeur in enumerate(lignes[0])}
    manquantes = [libelle for cle, libelle in _COLONNES_OBLIGATOIRES.items() if cle not in index_colonnes]
    if manquantes:
        return ResultatImportMateriel(erreurs=[
            f"Colonnes manquantes ou mal nommées : {', '.join(manquantes)}. "
            "Téléchargez le modèle pour connaître le format attendu."
        ])

    resultat = ResultatImportMateriel()
    profil = getattr(user, 'profile', None)
    numeros_serie_vus = set()
    for numero, ligne in enumerate(lignes[1:], start=2):
        if _ligne_vide(ligne):
            continue
        resultat.total_lignes += 1
        try:
            with transaction.atomic():
                _creer_asset_depuis_ligne(ligne, index_colonnes, user, profil, numeros_serie_vus)
            resultat.crees += 1
        except _ErreurLigneImport as exc:
            resultat.erreurs.append(f"Ligne {numero} : {exc}")
        except ValidationError as exc:
            resultat.erreurs.append(f"Ligne {numero} : {'; '.join(exc.messages)}")
    return resultat
