"""Export tableur (CSV/XLSX) des listes Matériels et Installations.

Vérifie surtout que l'export respecte STRICTEMENT le périmètre de
l'utilisateur connecté, y compris quand celui-ci tente de contourner le
filtrage en manipulant les paramètres de la requête (ship/service/sector/
section) — l'export ne doit jamais exposer de données hors périmètre.
"""
import io

import openpyxl
from django.contrib.auth.models import User
from django.test import TestCase

from accounts.models import UserProfile
from assets.models import Asset, AssetType, Installation
from org.models import Sector, Service, Ship


class ExportAssetsTests(TestCase):
    def setUp(self):
        self.navire = Ship.objects.create(name="Navire export")
        self.service = Service.objects.create(ship=self.navire, name="Service export")
        self.secteur = Sector.objects.create(service=self.service, name="Secteur export")
        self.autre_secteur = Sector.objects.create(service=self.service, name="Autre secteur")
        self.type_materiel = AssetType.objects.create(
            name="Extincteur", category="Sécurité incendie", sector=self.secteur
        )

        self.chef = User.objects.create_user(username="chef_export", password="pass")
        UserProfile.objects.filter(user=self.chef).update(role="CHEF_SECTEUR", sector=self.secteur)
        self.chef.refresh_from_db()

        self.asset_perimetre = Asset.objects.create(
            asset_type=self.type_materiel, designation="Extincteur coursive",
            internal_id="EXT-001", ship=self.navire, service=self.service, sector=self.secteur,
        )
        self.asset_hors_perimetre = Asset.objects.create(
            asset_type=self.type_materiel, designation="Extincteur passerelle",
            internal_id="EXT-999", ship=self.navire, service=self.service, sector=self.autre_secteur,
        )

    def test_export_csv_respecte_le_perimetre(self):
        self.client.login(username="chef_export", password="pass")
        response = self.client.get("/assets/", {"export": "csv"})
        self.assertEqual(response.status_code, 200)
        contenu = response.content.decode("utf-8-sig")
        self.assertIn("Extincteur coursive", contenu)
        self.assertNotIn("Extincteur passerelle", contenu)

    def test_export_ignore_une_tentative_de_contournement_par_filtre(self):
        """Même en demandant explicitement le secteur voisin via le filtre GET,
        l'export ne doit renvoyer que le périmètre propre de l'utilisateur."""
        self.client.login(username="chef_export", password="pass")
        response = self.client.get(
            "/assets/", {"export": "csv", "sector": str(self.autre_secteur.id)}
        )
        contenu = response.content.decode("utf-8-sig")
        self.assertNotIn("Extincteur passerelle", contenu)

    def test_export_csv_contenu_et_entetes(self):
        self.client.login(username="chef_export", password="pass")
        response = self.client.get("/assets/", {"export": "csv"})
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertIn("attachment", response["Content-Disposition"])
        lignes = response.content.decode("utf-8-sig").splitlines()
        self.assertIn("Désignation", lignes[0])
        self.assertTrue(any("EXT-001" in ligne for ligne in lignes))

    def test_export_xlsx_format_valide(self):
        self.client.login(username="chef_export", password="pass")
        response = self.client.get("/assets/", {"export": "xlsx"})
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        classeur = openpyxl.load_workbook(io.BytesIO(response.content))
        feuille = classeur.active
        valeurs = [cell.value for row in feuille.iter_rows() for cell in row]
        self.assertIn("Extincteur coursive", valeurs)
        self.assertNotIn("Extincteur passerelle", valeurs)


class ExportInstallationsTests(TestCase):
    def setUp(self):
        self.navire = Ship.objects.create(name="Navire export inst")
        self.service = Service.objects.create(ship=self.navire, name="Service export inst")
        self.secteur = Sector.objects.create(service=self.service, name="Secteur export inst")
        self.autre_secteur = Sector.objects.create(service=self.service, name="Autre secteur inst")

        self.chef = User.objects.create_user(username="chef_export_inst", password="pass")
        UserProfile.objects.filter(user=self.chef).update(role="CHEF_SECTEUR", sector=self.secteur)
        self.chef.refresh_from_db()

        self.installation_perimetre = Installation.objects.create(
            designation="Pompe principale", ship=self.navire, service=self.service, sector=self.secteur,
        )
        self.installation_hors_perimetre = Installation.objects.create(
            designation="Pompe voisine", ship=self.navire, service=self.service, sector=self.autre_secteur,
        )

    def test_export_csv_respecte_le_perimetre(self):
        self.client.login(username="chef_export_inst", password="pass")
        response = self.client.get("/installations/", {"export": "csv"})
        contenu = response.content.decode("utf-8-sig")
        self.assertIn("Pompe principale", contenu)
        self.assertNotIn("Pompe voisine", contenu)

    def test_export_ignore_une_tentative_de_contournement_par_filtre(self):
        self.client.login(username="chef_export_inst", password="pass")
        response = self.client.get(
            "/installations/", {"export": "csv", "sector": str(self.autre_secteur.id)}
        )
        contenu = response.content.decode("utf-8-sig")
        self.assertNotIn("Pompe voisine", contenu)

    def test_export_xlsx_format_valide(self):
        self.client.login(username="chef_export_inst", password="pass")
        response = self.client.get("/installations/", {"export": "xlsx"})
        classeur = openpyxl.load_workbook(io.BytesIO(response.content))
        feuille = classeur.active
        valeurs = [cell.value for row in feuille.iter_rows() for cell in row]
        self.assertIn("Pompe principale", valeurs)
        self.assertNotIn("Pompe voisine", valeurs)
