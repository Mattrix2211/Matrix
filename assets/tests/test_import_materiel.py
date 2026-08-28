"""Import en masse de matériel mobile depuis un fichier Excel (Phase 6).

Vérifie le cas nominal, l'import atomique par ligne (une ligne en erreur au
milieu d'un lot valide n'empêche pas les autres), le respect du périmètre, et
les cas d'erreurs de fichier (vide, colonnes manquantes).
"""
import io

import openpyxl
from django.contrib.auth.models import User
from django.test import TestCase

from accounts.models import UserProfile
from assets.models import Asset, AssetType
from assets.import_materiel import ENTETES_MODELE, importer_materiel_depuis_fichier
from org.models import Sector, Service, Ship


def _classeur_depuis_lignes(lignes):
    """Construit un fichier Excel en mémoire (comme le ferait un import HTTP réel)
    à partir d'une liste de lignes (la première étant les en-têtes)."""
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    for ligne in lignes:
        feuille.append(ligne)
    tampon = io.BytesIO()
    classeur.save(tampon)
    tampon.seek(0)
    tampon.name = "import.xlsx"
    return tampon


class ImportMaterielTests(TestCase):
    def setUp(self):
        self.navire = Ship.objects.create(name="Navire import")
        self.service = Service.objects.create(ship=self.navire, name="Service import")
        self.secteur = Sector.objects.create(service=self.service, name="Secteur import")
        self.autre_secteur = Sector.objects.create(service=self.service, name="Autre secteur import")
        self.type_extincteur = AssetType.objects.create(
            name="Extincteur", category="Sécurité incendie", sector=self.secteur
        )

        self.chef = User.objects.create_user(username="chef_import", password="pass")
        UserProfile.objects.filter(user=self.chef).update(role="CHEF_SECTEUR", sector=self.secteur)
        self.chef.refresh_from_db()

        self.equipier = User.objects.create_user(username="equipier_import", password="pass")
        UserProfile.objects.filter(user=self.equipier).update(role="EQUIPIER", sector=self.secteur)
        self.equipier.refresh_from_db()

    def _ligne(self, designation="Extincteur CO2", type_="Extincteur", **kwargs):
        valeurs = {
            'Désignation': designation, 'Type': type_, 'Identifiant interne': '',
            'N° série': '', 'Statut': '', 'Criticité': '', 'Unité': '', 'Service': '',
            'Secteur': '', 'Section': '', 'Emplacement': '', 'Référence': '', 'Marque': '',
            'NNO': '', 'Gisement': '', 'Local': '',
        }
        valeurs.update(kwargs)
        return [valeurs[entete] for entete in ENTETES_MODELE]

    def test_import_valide_cree_les_materiels(self):
        fichier = _classeur_depuis_lignes([
            ENTETES_MODELE,
            self._ligne("Extincteur coursive", "Extincteur", **{'N° série': 'SN-001'}),
            self._ligne("Extincteur passerelle", "Extincteur", **{'N° série': 'SN-002'}),
        ])
        resultat = importer_materiel_depuis_fichier(fichier, self.chef)
        self.assertEqual(resultat.total_lignes, 2)
        self.assertEqual(resultat.crees, 2)
        self.assertEqual(resultat.erreurs, [])
        self.assertEqual(Asset.objects.filter(sector=self.secteur).count(), 2)
        cree = Asset.objects.get(serial_number='SN-001')
        self.assertEqual(cree.ship_id, self.navire.id)
        self.assertEqual(cree.service_id, self.service.id)
        self.assertEqual(cree.asset_type_id, self.type_extincteur.id)

    def test_ligne_en_erreur_n_empeche_pas_les_autres(self):
        """Import atomique par ligne : une ligne avec un type de matériel inconnu
        au milieu d'un lot valide n'empêche pas les lignes correctes d'être créées."""
        fichier = _classeur_depuis_lignes([
            ENTETES_MODELE,
            self._ligne("Extincteur A", "Extincteur", **{'N° série': 'SN-A'}),
            self._ligne("Matériel inconnu", "TypeInexistant"),
            self._ligne("Extincteur B", "Extincteur", **{'N° série': 'SN-B'}),
        ])
        resultat = importer_materiel_depuis_fichier(fichier, self.chef)
        self.assertEqual(resultat.total_lignes, 3)
        self.assertEqual(resultat.crees, 2)
        self.assertEqual(len(resultat.erreurs), 1)
        self.assertIn("Ligne 3", resultat.erreurs[0])
        self.assertIn("type de matériel inconnu", resultat.erreurs[0])
        self.assertTrue(Asset.objects.filter(serial_number='SN-A').exists())
        self.assertTrue(Asset.objects.filter(serial_number='SN-B').exists())

    def test_numero_de_serie_deja_utilise_ou_duplique_est_accepte(self):
        """Un même numéro de série peut légitimement appartenir à plusieurs
        matériels différents (numérotations réutilisées entre fournisseurs) :
        ni un doublon avec un matériel déjà existant, ni un doublon entre deux
        lignes du même fichier, ne doit être rejeté."""
        Asset.objects.create(
            asset_type=self.type_extincteur, designation="Existant", serial_number="SN-DEJA",
            ship=self.navire, service=self.service, sector=self.secteur,
        )
        fichier = _classeur_depuis_lignes([
            ENTETES_MODELE,
            self._ligne("Nouveau", "Extincteur", **{'N° série': 'SN-DEJA'}),
            self._ligne("Autre nouveau", "Extincteur", **{'N° série': 'SN-DEJA'}),
        ])
        resultat = importer_materiel_depuis_fichier(fichier, self.chef)
        self.assertEqual(resultat.crees, 2)
        self.assertEqual(resultat.erreurs, [])
        self.assertEqual(Asset.objects.filter(serial_number='SN-DEJA').count(), 3)

    def test_designation_manquante_est_signalee(self):
        fichier = _classeur_depuis_lignes([
            ENTETES_MODELE,
            self._ligne(designation=""),
        ])
        resultat = importer_materiel_depuis_fichier(fichier, self.chef)
        self.assertEqual(resultat.crees, 0)
        self.assertIn("Ligne 2", resultat.erreurs[0])
        self.assertIn("désignation obligatoire manquante", resultat.erreurs[0])

    def test_fichier_vide_est_signale(self):
        fichier = _classeur_depuis_lignes([])
        resultat = importer_materiel_depuis_fichier(fichier, self.chef)
        self.assertEqual(resultat.crees, 0)
        self.assertIn("vide", resultat.erreurs[0])

    def test_colonnes_manquantes_sont_signalees(self):
        fichier = _classeur_depuis_lignes([
            ["Nom du matériel", "Catégorie technique"],
            ["Extincteur", "Sécurité"],
        ])
        resultat = importer_materiel_depuis_fichier(fichier, self.chef)
        self.assertEqual(resultat.crees, 0)
        self.assertIn("Colonnes manquantes", resultat.erreurs[0])

    def test_secteur_hors_perimetre_est_refuse(self):
        fichier = _classeur_depuis_lignes([
            ENTETES_MODELE,
            self._ligne("Extincteur voisin", "Extincteur", Secteur=self.autre_secteur.name),
        ])
        resultat = importer_materiel_depuis_fichier(fichier, self.chef)
        self.assertEqual(resultat.crees, 0)
        self.assertIn("hors de votre périmètre", resultat.erreurs[0])

    def test_lignes_vides_sont_ignorees_sans_compter_dans_le_total(self):
        fichier = _classeur_depuis_lignes([
            ENTETES_MODELE,
            self._ligne("Extincteur seul", "Extincteur", **{'N° série': 'SN-SEUL'}),
            [None] * len(ENTETES_MODELE),
        ])
        resultat = importer_materiel_depuis_fichier(fichier, self.chef)
        self.assertEqual(resultat.total_lignes, 1)
        self.assertEqual(resultat.crees, 1)


class ImportMaterielVueTests(TestCase):
    """Vérifie l'accès à l'écran d'import (rôle requis) et le téléchargement du
    modèle Excel, en plus du traitement métier déjà couvert ci-dessus."""

    def setUp(self):
        self.navire = Ship.objects.create(name="Navire import vue")
        self.service = Service.objects.create(ship=self.navire, name="Service import vue")
        self.secteur = Sector.objects.create(service=self.service, name="Secteur import vue")
        self.chef = User.objects.create_user(username="chef_import_vue", password="pass")
        UserProfile.objects.filter(user=self.chef).update(role="CHEF_SECTION", sector=self.secteur)
        self.chef.refresh_from_db()
        self.equipier = User.objects.create_user(username="equipier_import_vue", password="pass")
        UserProfile.objects.filter(user=self.equipier).update(role="EQUIPIER", sector=self.secteur)
        self.equipier.refresh_from_db()

    def test_equipier_ne_peut_pas_acceder_a_l_import(self):
        self.client.login(username="equipier_import_vue", password="pass")
        response = self.client.get("/assets/importer/")
        self.assertEqual(response.status_code, 403)

    def test_chef_peut_telecharger_le_modele(self):
        self.client.login(username="chef_import_vue", password="pass")
        response = self.client.get("/assets/importer/modele/")
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        classeur = openpyxl.load_workbook(io.BytesIO(response.content))
        feuille = classeur["Modèle"]
        entetes = [cell.value for cell in next(feuille.iter_rows())]
        self.assertEqual(entetes, ENTETES_MODELE)

    def test_import_via_la_vue_affiche_le_rapport(self):
        AssetType.objects.create(name="Extincteur", category="Sécurité", sector=self.secteur)
        classeur = openpyxl.Workbook()
        feuille = classeur.active
        feuille.append(ENTETES_MODELE)
        feuille.append(["Extincteur test", "Extincteur"] + [""] * (len(ENTETES_MODELE) - 2))
        tampon = io.BytesIO()
        classeur.save(tampon)
        tampon.seek(0)
        tampon.name = "import.xlsx"

        self.client.login(username="chef_import_vue", password="pass")
        response = self.client.post("/assets/importer/", {"fichier": tampon})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "1 matériel(s) créé(s)")
        self.assertTrue(Asset.objects.filter(designation="Extincteur test").exists())
