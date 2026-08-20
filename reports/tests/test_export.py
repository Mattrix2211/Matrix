"""Export tableur (CSV/XLSX) des bilans — même contexte de données que le PDF
(reports/services.py), juste un rendu différent.

Vérifie le contenu du fichier généré et que le périmètre reste strictement
respecté (PerimetreNonAutorise, même garde-fou que le PDF), à la fois côté
service et côté vue web (paramètre `format`).
"""
import io

import openpyxl
from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from accounts.models import UserProfile
from assets.models import Installation
from org.models import Sector, Service, Ship
from reports.services import (
    PerimetreNonAutorise,
    generer_bilan_instantane_csv,
    generer_bilan_instantane_xlsx,
    generer_bilan_periode_csv,
    generer_bilan_periode_xlsx,
)


class BilanExportServiceTests(TestCase):
    def setUp(self):
        self.navire = Ship.objects.create(name="Navire export bilan", code="EXB")
        self.service = Service.objects.create(ship=self.navire, name="Service export bilan")
        self.secteur = Sector.objects.create(service=self.service, name="Secteur export bilan")
        self.autre_secteur = Sector.objects.create(service=self.service, name="Autre secteur bilan")

        self.chef = User.objects.create_user(username="chef_bilan", password="pass")
        UserProfile.objects.filter(user=self.chef).update(role="CHEF_SECTEUR", sector=self.secteur)
        self.chef.refresh_from_db()

        self.installation = Installation.objects.create(
            designation="Pompe principale", ship=self.navire, service=self.service, sector=self.secteur,
        )
        self.installation_hors_perimetre = Installation.objects.create(
            designation="Pompe voisine", ship=self.navire, service=self.service, sector=self.autre_secteur,
        )

    def test_csv_instantane_respecte_le_perimetre(self):
        contenu = generer_bilan_instantane_csv("sector", self.secteur.id, self.chef)
        texte = contenu.decode("utf-8-sig")
        self.assertIn("Pompe principale", texte)
        self.assertNotIn("Pompe voisine", texte)

    def test_xlsx_instantane_respecte_le_perimetre(self):
        contenu = generer_bilan_instantane_xlsx("sector", self.secteur.id, self.chef)
        classeur = openpyxl.load_workbook(io.BytesIO(contenu))
        valeurs = [
            cell.value
            for feuille in classeur.worksheets
            for row in feuille.iter_rows()
            for cell in row
        ]
        self.assertIn("Pompe principale", valeurs)
        self.assertNotIn("Pompe voisine", valeurs)

    def test_csv_instantane_perimetre_non_autorise(self):
        with self.assertRaises(PerimetreNonAutorise):
            generer_bilan_instantane_csv("sector", self.autre_secteur.id, self.chef)

    def test_csv_periode_respecte_le_perimetre(self):
        aujourdhui = timezone.localdate()
        contenu = generer_bilan_periode_csv(
            "sector", self.secteur.id, self.chef, aujourdhui, aujourdhui
        )
        texte = contenu.decode("utf-8-sig")
        self.assertIn("Maintenances de la période", texte)

    def test_xlsx_periode_perimetre_non_autorise(self):
        aujourdhui = timezone.localdate()
        with self.assertRaises(PerimetreNonAutorise):
            generer_bilan_periode_xlsx(
                "sector", self.autre_secteur.id, self.chef, aujourdhui, aujourdhui
            )


class GenererBilanViewFormatTests(TestCase):
    def setUp(self):
        self.navire = Ship.objects.create(name="Navire export bilan vue", code="EBV")
        self.service = Service.objects.create(ship=self.navire, name="Service export bilan vue")
        self.secteur = Sector.objects.create(service=self.service, name="Secteur export bilan vue")

        self.chef = User.objects.create_user(username="chef_bilan_vue", password="pass")
        UserProfile.objects.filter(user=self.chef).update(role="CHEF_SECTEUR", sector=self.secteur)
        self.chef.refresh_from_db()

        Installation.objects.create(
            designation="Pompe vue", ship=self.navire, service=self.service, sector=self.secteur,
        )

        self.url = reverse("rapport-bilan")

    def test_format_csv_renvoie_un_csv_en_telechargement(self):
        self.client.login(username="chef_bilan_vue", password="pass")
        response = self.client.get(self.url, {"format": "csv"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn("Pompe vue", response.content.decode("utf-8-sig"))

    def test_format_xlsx_renvoie_un_classeur_valide(self):
        self.client.login(username="chef_bilan_vue", password="pass")
        response = self.client.get(self.url, {"format": "xlsx"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        # Vérifie simplement que le contenu se charge comme un classeur XLSX valide.
        openpyxl.load_workbook(io.BytesIO(response.content))

    def test_format_csv_en_mode_periode(self):
        self.client.login(username="chef_bilan_vue", password="pass")
        aujourdhui = timezone.localdate()
        response = self.client.get(
            self.url,
            {"format": "csv", "date_debut": aujourdhui.isoformat(), "date_fin": aujourdhui.isoformat()},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
