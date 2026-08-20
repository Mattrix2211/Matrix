"""Utilitaires génériques d'export tableur (CSV / XLSX).

Point de rendu unique réutilisé par les vues liste (matériels, installations,
stock de pièces) et par les bilans (reports/services.py), pour ne jamais
dupliquer la construction du fichier : chaque appelant fournit seulement les
en-têtes de colonnes et les lignes déjà préparées (valeurs déjà mises en forme
en français), ce module se charge uniquement du rendu CSV ou XLSX.

XLSX est une dépendance optionnelle (openpyxl, bibliothèque pure Python sans
dépendance CDN, cf. requirements.txt) : si elle n'est pas installée, seul le
CSV reste disponible (aucune dépendance pour ce format, module `csv` du socle
Python).
"""
import csv
import io

from django.http import HttpResponse

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover - openpyxl est listée dans requirements.txt
    Workbook = None

CSV_CONTENT_TYPE = "text/csv"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Séparateur ';' plutôt que ',' : c'est le séparateur par défaut d'Excel en
# configuration régionale française (sinon toutes les colonnes atterrissent
# dans une seule cellule à l'ouverture du fichier).
_SEPARATEUR_CSV = ";"


def xlsx_disponible() -> bool:
    """Indique si le rendu XLSX est disponible (openpyxl installé)."""
    return Workbook is not None


def rendre_csv(entetes, lignes) -> bytes:
    """Construit un CSV à partir d'une liste d'en-têtes et de lignes (une seule
    section). Encodage UTF-8 avec BOM pour qu'Excel affiche correctement les
    caractères accentués à l'ouverture."""
    tampon = io.StringIO()
    writer = csv.writer(tampon, delimiter=_SEPARATEUR_CSV)
    writer.writerow(entetes)
    writer.writerows(lignes)
    return ("﻿" + tampon.getvalue()).encode("utf-8")


def rendre_csv_sections(sections) -> bytes:
    """Construit un CSV à partir de plusieurs sections (titre, en-têtes, lignes),
    empilées et séparées par une ligne vide — utilisé pour les bilans, qui
    couvrent plusieurs types de données (installations, tickets, formations...)
    dans un fichier CSV ne supportant qu'une seule feuille."""
    tampon = io.StringIO()
    writer = csv.writer(tampon, delimiter=_SEPARATEUR_CSV)
    for index, (titre, entetes, lignes) in enumerate(sections):
        if index > 0:
            writer.writerow([])
        writer.writerow([titre])
        writer.writerow(entetes)
        writer.writerows(lignes)
    return ("﻿" + tampon.getvalue()).encode("utf-8")


def rendre_xlsx(entetes, lignes, titre_feuille="Export") -> bytes | None:
    """Construit un classeur XLSX à une feuille. Renvoie None si openpyxl n'est
    pas installé (l'appelant doit alors se rabattre sur le CSV)."""
    if Workbook is None:
        return None
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = titre_feuille[:31]  # limite Excel : 31 caractères max
    feuille.append(list(entetes))
    for ligne in lignes:
        feuille.append(list(ligne))
    tampon = io.BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()


def rendre_xlsx_sections(sections) -> bytes | None:
    """Construit un classeur XLSX avec une feuille par section (titre, en-têtes,
    lignes) — utilisé pour les bilans. Renvoie None si openpyxl n'est pas
    installé."""
    if Workbook is None:
        return None
    classeur = Workbook()
    classeur.remove(classeur.active)
    for titre, entetes, lignes in sections:
        feuille = classeur.create_sheet(title=titre[:31])
        feuille.append(list(entetes))
        for ligne in lignes:
            feuille.append(list(ligne))
    tampon = io.BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()


def reponse_fichier(contenu: bytes, nom_fichier: str, content_type: str) -> HttpResponse:
    """Construit la réponse HTTP de téléchargement du fichier exporté."""
    reponse = HttpResponse(contenu, content_type=content_type)
    reponse["Content-Disposition"] = f'attachment; filename="{nom_fichier}"'
    return reponse


def construire_url_export(request, format_export: str) -> str:
    """Construit l'URL d'export (CSV ou XLSX) en conservant les filtres actifs
    de la page courante (querystring), pour que l'export corresponde toujours
    à ce que l'utilisateur a sous les yeux au moment du clic."""
    parametres = request.GET.copy()
    parametres["export"] = format_export
    return f"{request.path}?{parametres.urlencode()}"
